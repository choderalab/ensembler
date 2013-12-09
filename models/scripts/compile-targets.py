# Compile sequence data for comparative modeling.
#
# John D. Chodera <choderaj@mskcc.org> - 15 Jan 2013
#
# Based on:
# A sample script for fully automated comparative modeling (Ben Webb)
#
# PREREQUISITES
#
# * openpyxl 
# http://packages.python.org/openpyxl/
#
# TODO
# 
# * Create a separate directory for each target sequence (with separate exceptions.out and sequence identity files)
# * Split out template sequence extraction into a separate script.

# PARAMETERS

# Reject sequences below this length.
min_sequence_length = 10

# Print verbose debugging information during execution.
verbose = True 

#
# GET ABSOLUTE PATHS
#

import os.path

# Input directories.
sequence_directory = os.path.abspath("../sequences")

# Output directories.
targets_directory = os.path.abspath("targets")

# Output files.
import os.path
targets_index_filename = os.path.join(targets_directory, 'targets.txt') # list of target names
targets_sequences_filename = os.path.join(targets_directory, 'targets.seg') # sequence input file for MODELLER

# 
# CREATE TARGETS DIRECTORY
#

import os
if not os.path.exists(targets_directory):
    os.mkdir(targets_directory)

#
# BUILD KINASE DOMAIN SEQUENCE LIST
#

contents = "C; Alignment file\n\n" # alignment file
targets = list()

# Build a list of sequences.
from openpyxl import load_workbook
workbook_filename = os.path.join(sequence_directory, 'kinbase', 'Kincat_Hsap.08.02.xlsx')
wb = load_workbook(filename = workbook_filename)
sheet_ranges = wb.get_sheet_by_name(name = 'Sheet1')
nrows = sheet_ranges.get_highest_row()
for row in range(2,nrows):
    name = str(sheet_ranges.cell('A%d' % row).value) # name
    pseudogene = sheet_ranges.cell('F%d' % row).value # ['Y','R','N']
    sequence = str(sheet_ranges.cell('J%d' % row).value)     

    # Reject sequences that are too short.
    if len(sequence) < min_sequence_length: continue

    # Ignore pseudogenes.
    if pseudogene != 'N': continue

    # Write alignment file.
    contents +=  ">P1;%s\n" % name
    contents += "sequence:%s:FIRST:@:LAST:@::::\n" % name
    contents += sequence + "*\n"    

    # Write output.
    if verbose:
        print "row %4d : %12s : %5d : %s" % (row, name, len(sequence), sequence)

    # Append to list of targets.
    targets.append( name )

outfile = open(targets_sequences_filename, 'w')
outfile.write(contents)
outfile.close()

#
# BUILD TARGETS INDEX FILE
#

contents = ""
for target in targets:
    contents += "%s\n" % target
    
outfile = open(targets_index_filename, 'w')
outfile.write(contents)
outfile.close()

if verbose: print "\n%d target sequences written." % len(targets)

